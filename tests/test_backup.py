"""
Unit tests for backup.py — encrypted ZIP export/import, Excel export, and
the on-disk backup history used by the Admin portal.

backup.py imports several path constants and functions by value from
portfolio.py / settings.py at module load time, so isolation requires
patching backup.py's own bindings (_SHARED_FILE, _ENV_FILE, _BACKUPS_DIR,
_BACKUPS_MANIFEST) in addition to portfolio._BASE_DIR / settings._DATA_DIR.
"""

import io
import zipfile

import openpyxl
import pandas as pd
import pytest

import backup
import portfolio
import settings


@pytest.fixture(autouse=True)
def isolated_backup(tmp_path, monkeypatch):
    monkeypatch.setenv("ENCRYPTION_KEY", "unit-test-key-123")
    monkeypatch.setattr(portfolio, "_BASE_DIR", tmp_path / "portfolio")
    monkeypatch.setattr(settings, "_DATA_DIR", tmp_path / "settings_data")
    monkeypatch.setattr(settings, "_SHARED_FILE", tmp_path / "settings_data" / "shared.json")
    monkeypatch.setattr(backup, "_SHARED_FILE", tmp_path / "settings_data" / "shared.json")
    monkeypatch.setattr(backup, "_ENV_FILE", tmp_path / "fake.env")
    monkeypatch.setattr(backup, "_BACKUPS_DIR", tmp_path / "backups")
    monkeypatch.setattr(backup, "_BACKUPS_MANIFEST", tmp_path / "backups" / "manifest.json")
    portfolio.set_user("test@example.com")
    yield
    portfolio.set_user("")


EMAIL = "test@example.com"


def _zip_names(zip_bytes: bytes) -> set[str]:
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        return set(zf.namelist())


# ── export_zip ────────────────────────────────────────────────────────────

class TestExportZip:
    def test_includes_only_existing_portfolio_files(self):
        portfolio.save_portfolio(pd.DataFrame([{"ticker": "AAA.BR"}]))
        names = _zip_names(backup.export_zip(EMAIL))
        assert "data/portfolio.json" in names
        assert "data/sold.json" not in names

    def test_no_data_produces_empty_but_valid_zip(self):
        names = _zip_names(backup.export_zip(EMAIL))
        assert names == set()

    def test_includes_env_file_when_present(self):
        backup._ENV_FILE.write_text("AUTH_SECRET=abc\n")
        names = _zip_names(backup.export_zip(EMAIL))
        assert ".env" in names

    def test_includes_shared_settings_when_present(self):
        settings.save_shared_settings({"max_debt_equity": 400.0})
        names = _zip_names(backup.export_zip(EMAIL))
        assert "data/shared_settings.json" in names

    def test_includes_user_settings_when_email_given(self):
        settings.save_settings({"density": "compact"}, EMAIL)
        names = _zip_names(backup.export_zip(EMAIL))
        assert "data/settings.json" in names

    def test_omits_user_settings_when_no_email(self):
        settings.save_settings({"density": "compact"}, EMAIL)
        names = _zip_names(backup.export_zip(""))
        assert "data/settings.json" not in names


# ── export_excel ──────────────────────────────────────────────────────────

class TestExportExcel:
    def test_raises_when_no_data(self):
        with pytest.raises(ValueError):
            backup.export_excel()

    def test_produces_expected_sheets(self):
        portfolio.save_portfolio(pd.DataFrame([{"ticker": "AAA.BR"}]))
        portfolio.save_sold(pd.DataFrame([{"ticker": "BBB.BR"}]))
        portfolio.save_div_hist(pd.DataFrame([{"ticker": "AAA.BR", "amount": 5.0}]))
        portfolio.save_watchlist({"CCC.BR"})

        wb = openpyxl.load_workbook(io.BytesIO(backup.export_excel()))
        assert set(wb.sheetnames) == {"Positions", "Sold", "Dividends", "Watchlist"}

    def test_omits_sheets_for_empty_sections(self):
        portfolio.save_portfolio(pd.DataFrame([{"ticker": "AAA.BR"}]))
        wb = openpyxl.load_workbook(io.BytesIO(backup.export_excel()))
        assert wb.sheetnames == ["Positions"]


# ── import_zip ────────────────────────────────────────────────────────────

class TestImportZip:
    def _make_zip(self, entries: dict[str, bytes]) -> bytes:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            for name, content in entries.items():
                zf.writestr(name, content)
        return buf.getvalue()

    def test_restores_portfolio_files(self):
        zip_bytes = self._make_zip({"data/portfolio.json": b'[{"ticker": "AAA.BR"}]'})
        restored = backup.import_zip(zip_bytes, EMAIL)
        assert "portfolio.json" in restored
        assert (portfolio.user_data_dir(EMAIL) / "portfolio.json").read_bytes() == b'[{"ticker": "AAA.BR"}]'

    def test_ignores_data_files_not_in_allowlist(self):
        zip_bytes = self._make_zip({"data/not_recognised.json": b"{}", "data/portfolio.json": b"[]"})
        restored = backup.import_zip(zip_bytes, EMAIL)
        assert restored == ["portfolio.json"]

    def test_restores_env_file(self):
        zip_bytes = self._make_zip({".env": b"AUTH_SECRET=xyz\n"})
        restored = backup.import_zip(zip_bytes, EMAIL)
        assert ".env" in restored
        assert backup._ENV_FILE.read_bytes() == b"AUTH_SECRET=xyz\n"

    def test_restores_shared_settings(self):
        zip_bytes = self._make_zip({"data/shared_settings.json": b'{"max_debt_equity": 300.0}'})
        restored = backup.import_zip(zip_bytes, EMAIL)
        assert "shared settings" in restored
        assert backup._SHARED_FILE.read_bytes() == b'{"max_debt_equity": 300.0}'

    def test_restores_user_settings_when_email_given(self):
        zip_bytes = self._make_zip({"data/settings.json": b'{"density": "compact"}'})
        restored = backup.import_zip(zip_bytes, EMAIL)
        assert "settings" in restored
        assert settings._settings_file(EMAIL).read_bytes() == b'{"density": "compact"}'

    def test_skips_user_settings_when_no_email(self):
        zip_bytes = self._make_zip({"data/settings.json": b'{"density": "compact"}'})
        restored = backup.import_zip(zip_bytes, "")
        assert restored == []

    def test_rejects_invalid_zip_bytes(self):
        with pytest.raises(ValueError):
            backup.import_zip(b"this is not a zip file", EMAIL)

    def test_rejects_zip_without_recognisable_backup_content(self):
        zip_bytes = self._make_zip({"readme.txt": b"hello"})
        with pytest.raises(ValueError):
            backup.import_zip(zip_bytes, EMAIL)


class TestBackupFilename:
    def test_format_matches_expected_pattern(self):
        name = backup.backup_filename("zip")
        assert name.startswith("uv_backup_")
        assert name.endswith(".zip")
        # uv_backup_YYYYMMDD_HHMM.zip
        stamp = name[len("uv_backup_"):-len(".zip")]
        assert len(stamp) == len("20240101_1200")


# ── backup history (Admin portal) ────────────────────────────────────────

class TestBackupHistory:
    def test_create_backup_writes_zip_and_manifest_entry(self):
        portfolio.save_portfolio(pd.DataFrame([{"ticker": "AAA.BR"}]))
        entry = backup.create_backup(EMAIL)
        assert entry["email"] == EMAIL
        assert entry["type"] == "Manual"
        assert (backup._BACKUPS_DIR / entry["filename"]).exists()
        assert backup.list_backups() == [entry]

    def test_list_backups_sorted_newest_first(self):
        first = backup.create_backup(EMAIL)
        first["created_at"] = "2000-01-01T00:00:00+00:00"
        entries = backup._load_backup_manifest()
        entries[0] = first
        backup._save_backup_manifest(entries)

        second = backup.create_backup(EMAIL)
        listed = backup.list_backups()
        assert listed[0]["id"] == second["id"]
        assert listed[1]["id"] == first["id"]

    def test_get_backup_bytes_returns_zip_contents(self):
        portfolio.save_portfolio(pd.DataFrame([{"ticker": "AAA.BR"}]))
        entry = backup.create_backup(EMAIL)
        zip_bytes = backup.get_backup_bytes(entry["id"])
        assert "data/portfolio.json" in _zip_names(zip_bytes)

    def test_get_backup_bytes_raises_for_unknown_id(self):
        with pytest.raises(ValueError):
            backup.get_backup_bytes("does-not-exist")

    def test_get_backup_bytes_raises_when_file_missing_from_disk(self):
        entry = backup.create_backup(EMAIL)
        (backup._BACKUPS_DIR / entry["filename"]).unlink()
        with pytest.raises(ValueError):
            backup.get_backup_bytes(entry["id"])

    def test_restore_backup_roundtrips(self):
        portfolio.save_portfolio(pd.DataFrame([{"ticker": "AAA.BR"}]))
        entry = backup.create_backup(EMAIL)

        portfolio.save_portfolio(pd.DataFrame([{"ticker": "CHANGED.BR"}]))
        restored = backup.restore_backup(entry["id"], EMAIL)

        assert "portfolio.json" in restored
        assert portfolio.load_portfolio().iloc[0]["ticker"] == "AAA.BR"

    def test_load_backup_manifest_returns_empty_list_when_missing(self):
        assert backup._load_backup_manifest() == []

    def test_load_backup_manifest_returns_empty_list_on_corrupt_file(self):
        from crypto import write_encrypted
        backup._BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
        write_encrypted(backup._BACKUPS_MANIFEST, "not valid json{{{")
        assert backup._load_backup_manifest() == []
